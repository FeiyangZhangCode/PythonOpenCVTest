import math
import time
import smbus
import datetime
import openpyxl


# 电源管理寄存器地址

power_mgmt_1 = 0x6b

power_mgmt_2 = 0x6c


def read_byte(adr):
    return bus.read_byte_data(address, adr)


def read_word(adr):
    high = bus.read_byte_data(address, adr)

    low = bus.read_byte_data(address, adr + 1)

    val = (high << 8) + low

    return val


def read_word_2c(adr):
    val = read_word(adr)

    if val >= 0x8000:

        return -((65535 - val) + 1)

    else:

        return val


def dist(a, b):
    return math.sqrt((a * a) + (b * b))

    # math.sqrt(x) 方法返回数字x的平方根


def get_y_rotation(x, y, z):
    radians = math.atan2(x, dist(y, z))
    # math.atan2(y, x) 返回给定的 X 及 Y 坐标值的反正切值
    return -math.degrees(radians)
    # math.degrees(x) 将弧度x转换为角度


def get_x_rotation(x, y, z):
    radians = math.atan2(y, dist(x, z))
    return math.degrees(radians)


bus = smbus.SMBus(0)  # or bus = smbus.SMBus(1) for Revision 2 boards

address = 0x68  # This is the address value read via the i2cdetect command

# Now wake the 6050 up as it starts in sleep mode

bus.write_byte_data(address, power_mgmt_1, 0)

kp = 100.0
ki = 0.002
half_t = 0.05
q0 = 1.0
q1 = 0.0
q2 = 0.0
q3 = 0.0
ex_int = 0.0
ey_int = 0.0
ez_int = 0.0

work_book = openpyxl.load_workbook('./TestData/MPU.xlsx')
work_sheet = work_book['Sheet1']

while True:
    time.sleep(0.5)
    str_Time = datetime.datetime.now().strftime('%H:%M:%S.%f')

    temp_out = read_word_2c(0x41)
    temp_out = temp_out / 340 + 36.53
    # print("温度传感器")
    print("temperature: ", temp_out)

    gyro_xout = read_word_2c(0x43)
    gyro_yout = read_word_2c(0x45)
    gyro_zout = read_word_2c(0x47)
    gyro_xout_scaled = gyro_xout / 131
    gyro_yout_scaled = gyro_yout / 131
    gyro_zout_scaled = gyro_zout / 131
    # print("XYZ轴陀螺仪计数值,每秒的旋转度数")
    print("gyro_xout : ", gyro_xout, " scaled: ", gyro_xout_scaled)  # 倍率250/s
    print("gyro_yout : ", gyro_yout, " scaled: ", gyro_yout_scaled)
    print("gyro_zout : ", gyro_zout, " scaled: ", gyro_zout_scaled)

    accel_xout = read_word_2c(0x3b)
    accel_yout = read_word_2c(0x3d)
    accel_zout = read_word_2c(0x3f)
    accel_xout_scaled = accel_xout / 16384.0  # 倍率2g
    accel_yout_scaled = accel_yout / 16384.0
    accel_zout_scaled = accel_zout / 16384.0
    # print("XYZ轴加速度计数值,每秒的旋转度数")
    print("accel_xout: ", accel_xout, " scaled: ", accel_xout_scaled)
    print("accel_yout: ", accel_yout, " scaled: ", accel_yout_scaled)
    print("accel_zout: ", accel_zout, " scaled: ", accel_zout_scaled)

    # print("XY轴旋转度数")
    rot_x = get_x_rotation(accel_xout_scaled, accel_yout_scaled, accel_zout_scaled)
    rot_y = get_y_rotation(accel_xout_scaled, accel_yout_scaled, accel_zout_scaled)
    print("x rotation: ", rot_x)
    print("y rotation: ", rot_y)

    gx = gyro_xout_scaled
    gy = gyro_yout_scaled
    gz = gyro_zout_scaled
    ax = accel_xout_scaled
    ay = accel_yout_scaled
    az = accel_zout_scaled
    # 测量正常化
    norm = math.sqrt(ax * ax + ay * ay + az * az)
    ax = ax / norm
    ay = ay / norm
    az = az / norm
    # 估计方向的重力
    vx = 2 * (q1 * q3 - q0 * q2)
    vy = 2 * (q0 * q1 + q2 * q3)
    vz = q0 * q0 - q1 * q1 - q2 * q2 + q3 * q3
    # 错误的领域和方向传感器测量参考方向之间的交叉乘积的总和
    ex = (ay * vz - az * vy)
    ey = (az * vx - ax * vz)
    ez = (ax * vy - ay * vx)
    # 积分误差比例积分增益
    ex_int = ex_int + ex * ki
    ey_int = ey_int + ey * ki
    ez_int = ez_int + ez * ki
    # 调整后的陀螺仪测量
    gx = gx + kp * ex + ex_int
    gy = gy + kp * ey + ey_int
    gz = gz + kp * ez + ez_int
    # 整合四元数率和正常化
    q0 = q0 + (-q1 * gx - q2 * gy - q3 * gz) * half_t
    q1 = q1 + (q0 * gx + q2 * gz - q3 * gy) * half_t
    q2 = q2 + (q0 * gy - q1 * gz + q3 * gx) * half_t
    q3 = q3 + (q0 * gz + q1 * gy - q2 * gx) * half_t
    # 正常化四元
    norm = math.sqrt(q0 * q0 + q1 * q1 + q2 * q2 + q3 * q3)
    q0 = q0 / norm
    q1 = q1 / norm
    q2 = q2 / norm
    q3 = q3 / norm
    rot_pitch = math.asin(-2 * q1 * q3 + 2 * q0 * q2) * 57.3
    rot_roll = math.atan2(2 * q2 * q3 + 2 * q0 * q1, -2 * q1 * q1 - 2 * q2 * q2 + 1) * 57.3
    # print('四元欧拉角')
    print('Pitch: ', rot_pitch)
    print('Roll: ', rot_roll)

    # file_rec = open('./TestData/MPU.txt', 'a')
    # sav_mess = str_Time + '\n'
    # sav_mess += 'tmp:' + round(temp_out, 2) + '\n'
    # sav_mess += 'gxs:' + round(gyro_xout_scaled, 4) + '\n'
    # sav_mess += 'gys:' + round(gyro_yout_scaled, 4) + '\n'
    # sav_mess += 'gzs:' + round(gyro_zout_scaled, 4) + '\n'
    # sav_mess += 'axs:' + round(accel_xout_scaled, 4) + '\n'
    # sav_mess += 'ays:' + round(accel_yout_scaled, 4) + '\n'
    # sav_mess += 'azs:' + round(accel_zout_scaled, 4) + '\n'
    # sav_mess += 'xrt:' + round(rot_x, 4) + '\n'
    # sav_mess += 'yrt:' + round(rot_y, 4) + '\n'
    # sav_mess += 'pth:' + round(rot_pitch, 4) + '\n'
    # sav_mess += 'rol:' + round(rot_roll, 4) + '\n'
    # file_rec.write(sav_mess)
    # file_rec.close()
    row_num = work_sheet.max_row + 1
    work_sheet.cell(row_num, 1).value = str_Time
    work_sheet.cell(row_num, 2).value = round(temp_out, 2)
    work_sheet.cell(row_num, 3).value = round(gyro_xout_scaled, 4)
    work_sheet.cell(row_num, 4).value = round(gyro_yout_scaled, 4)
    work_sheet.cell(row_num, 5).value = round(gyro_zout_scaled, 4)
    work_sheet.cell(row_num, 6).value = round(accel_xout_scaled, 4)
    work_sheet.cell(row_num, 7).value = round(accel_yout_scaled, 4)
    work_sheet.cell(row_num, 8).value = round(accel_zout_scaled, 4)
    work_sheet.cell(row_num, 9).value = round(rot_x, 4)
    work_sheet.cell(row_num, 10).value = round(rot_y, 4)
    work_sheet.cell(row_num, 11).value = round(rot_pitch, 4)
    work_sheet.cell(row_num, 12).value = round(rot_roll, 4)
    work_book.save('./TestData/MPU.xlsx')



