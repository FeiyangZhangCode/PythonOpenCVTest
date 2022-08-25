import openpyxl
import datetime
import time

start_time = time.time()
work_book = openpyxl.load_workbook('./TestData/MPU.xlsx')
work_sheet = work_book['Sheet1']
# print(work_sheet.max_column)
# print(work_sheet.max_row)
# print(work_sheet.cell(2, 1).value)
row_num = work_sheet.max_row
work_sheet.cell(row_num + 1, 1).value = 'add00'
work_sheet.cell(row_num + 1, 2).value = 'add01'
work_book.save('./TestData/MPU.xlsx')
row_num = work_sheet.max_row
work_sheet.cell(row_num + 1, 1).value = 'add10'
work_sheet.cell(row_num + 1, 2).value = 'add11'
work_book.save('./TestData/MPU.xlsx')
end_time = time.time()
print(str(round((end_time - start_time) * 1000, 4)))
